import streamlit as st
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import pandas as pd

FILE_NAME = "money_tracker.xlsx"

# Create file if not exists
if not os.path.exists(FILE_NAME):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Daily_Expenses"
    ws1.append(["Date", "Item", "Quantity", "Price", "Total", "Money Left"])

    ws2 = wb.create_sheet(title="Summary")
    ws2["A1"] = "Total Spent"
    ws2["A2"] = "Wallet Balance"

    wb.save(FILE_NAME)

# Load workbook
wb = load_workbook(FILE_NAME)
ws1 = wb["Daily_Expenses"]
ws2 = wb["Summary"]

st.title("💰 Money Tracker")

# --- USER INPUT ---
item = st.text_input("Item")
quantity = st.number_input("Quantity", min_value=1, step=1)
price = st.number_input("Price per item", min_value=0.0, step=0.1)
wallet = st.number_input("Current wallet money", min_value=0.0, step=1.0)

if st.button("Add Expense"):
    date = datetime.now().strftime("%Y-%m-%d")
    total = quantity * price

    # Get last money left
    last_row = ws1.max_row
    if last_row > 1:
        last_money_left = ws1.cell(row=last_row, column=6).value
        if last_money_left is None:
            last_money_left = wallet
    else:
        last_money_left = wallet

    money_left = last_money_left - total

    # Save entry
    ws1.append([date, item, quantity, price, total, money_left])

    # Calculate total spent
    total_spent = 0
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row[4] is not None:
            total_spent += row[4]

    # Update summary
    ws2["B1"] = total_spent
    ws2["B2"] = money_left

    wb.save(FILE_NAME)

    st.success("Expense added!")

# --- DISPLAY DATA ---
data = ws1.values
columns = next(data)
df = pd.DataFrame(data, columns=columns)

st.subheader("📊 Daily Expenses")
st.dataframe(df)

# --- SUMMARY ---
total_spent = ws2["B1"].value
money_left = ws2["B2"].value

st.subheader("📈 Summary")
st.write(f"Total Spent: ${total_spent}")
st.write(f"Money Left: ${money_left}")
